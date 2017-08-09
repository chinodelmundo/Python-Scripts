#! python3
# amazonProductsScrape.py - Scrapes amazon products and reviews

import requests 
import os
import bs4
from openpyxl import Workbook
from openpyxl.styles import Font

# CONSTANTS
URL = "http://www.amazon.com"
SHEET1 = "Products"
SHEET2 = "Reviews"
PRODUCT_HEADERS = ["Product Name", "Price", "Rating", "Reviews", "Amazon Link"]
REVIEW_HEADERS = ["Author", "Review", "Rating", "Date"]

def main():
        wb = initialize_workbook()
        results = [];

        print("Enter product name: ")
        product = input()
        output_filename = str(product) + " - results.xlsx"

        print("Searching: " + product)
        res = requests.get(URL + "/s/?field-keywords=" + str(product))
        res.raise_for_status()

        soup = bs4.BeautifulSoup(res.text, "html.parser")
        items = soup.select("#s-results-list-atf .s-item-container")

        for item in items:
                print("Retrieving product details...")
                
                name = item.select(".s-access-detail-page h2")[0].getText()
                rating = item.select(".a-icon-star .a-icon-alt")[0].getText()
                link =  item.select(".s-access-detail-page")[0].attrs['href']

                # price
                currency = item.select(".sx-price-currency")[0].getText()
                whole = item.select(".sx-price-whole")[0].getText()
                decimal = item.select(".sx-price-fractional")[0].getText()
                price = currency + whole + "." + decimal

                item_details = {
                        "name" : name,
                        "rating" : rating,
                        "link" : link,
                        "price" : price
                }

                item_details["reviews"] = get_reviews(item_details["link"])
                wb = workbook_write(wb, item_details)
                workbook_save(wb, output_filename)

        print('Done! See excel file: ' + output_filename)


def initialize_workbook(): 
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET1
        wb.create_sheet(title=SHEET2)

        for i, header in enumerate(PRODUCT_HEADERS):
                ws.cell(row=1, column=i + 1).value = header
                ws.cell(row=1, column=i + 1).font = Font(bold=True)

        return wb

def get_reviews(product_link):
        print("Retrieving reviews...")
        reviews_list = []

        
        res = requests.get(product_link)
        res.raise_for_status()
        soup = bs4.BeautifulSoup(res.text, "html.parser")
        reviews_link = soup.select("#dp-summary-see-all-reviews")[0].attrs["href"]

        res = requests.get(URL + reviews_link)
        res.raise_for_status()
        soup = bs4.BeautifulSoup(res.text, "html.parser")
        reviews = soup.select("#cm_cr-review_list .review")

        for review in reviews:
                author = review.select(".author")[0].getText()
                text = review.select(".review-text")[0].getText()
                rating = review.select(".a-icon-star .a-icon-alt")[0].getText()
                date = review.select(".review-date")[0].getText()

                review_details = {
                        "author": author,
                        "text": text,
                        "rating": rating,
                        "date": date
                }

                reviews_list.append(review_details)

        return reviews_list

def workbook_write(wb, item_details):
        # Insert Product Details
        sheet1 = wb.get_sheet_by_name(SHEET1)
        sheet1_start_row = sheet1.max_row + 1
	
        sheet1.cell(row=sheet1_start_row, column=1).value = item_details["name"]
        sheet1.cell(row=sheet1_start_row, column=2).value = item_details["price"]
        sheet1.cell(row=sheet1_start_row, column=3).value = item_details["rating"]
        
        sheet1.cell(row=sheet1_start_row, column=4).value = "Reviews"
        sheet1.cell(row=sheet1_start_row, column=4).font = Font(underline='single', color='0563C1')
        
        sheet1.cell(row=sheet1_start_row, column=5).value = "Link"
        sheet1.cell(row=sheet1_start_row, column=5).font = Font(underline='single', color='0563C1')
        sheet1.cell(row=sheet1_start_row, column=5).hyperlink = item_details["link"]

        # Insert Reviews
        sheet2 = wb.get_sheet_by_name(SHEET2)            
        sheet2_start_row = sheet2.max_row
        if sheet2_start_row > 1:
                sheet2_start_row+= 3

        sheet1.cell(row=sheet1_start_row, column=4).hyperlink =  "#" + SHEET2 +"!A" + str(sheet2_start_row) #Set link to Reviews sheet

        sheet2.cell(row=sheet2_start_row, column=1).value = "Product Name:"
        sheet2.cell(row=sheet2_start_row, column=1).font = Font(bold=True)
        sheet2.cell(row=sheet2_start_row, column=2).value = item_details["name"]
                
        sheet2_start_row+= 1
        for i, header in enumerate(REVIEW_HEADERS):
                sheet2.cell(row=sheet2_start_row, column=i + 1).value = header
                sheet2.cell(row=sheet2_start_row, column=i + 1).font = Font(bold=True)

        sheet2_start_row+= 1
        for index, review in enumerate(item_details["reviews"]):
                sheet2.cell(row=sheet2_start_row + index, column=1).value = review["author"]
                sheet2.cell(row=sheet2_start_row + index, column=2).value = review["text"]
                sheet2.cell(row=sheet2_start_row + index, column=3).value = review["rating"]
                sheet2.cell(row=sheet2_start_row + index, column=4).value = review["date"]
        
        return wb

def workbook_save(wb, filename):
        print('+++++++++++++++++++++')
        print('Saving...')

        wb.save(filename)

        print('Done saving!')
        print('+++++++++++++++++++++')

if __name__ == '__main__':
        main()
